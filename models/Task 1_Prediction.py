"""Inference script for the Urban-ImageNet Task 1 scene classifier."""

from __future__ import annotations

import argparse
import csv
import os
from pathlib import Path
from typing import Iterable

os.environ.setdefault("KMP_DUPLICATE_LIB_OK", "TRUE")
os.environ.setdefault("OMP_NUM_THREADS", "1")

import torch
import torch.nn as nn
from PIL import Image
from torchvision import models, transforms

try:
    from tqdm import tqdm
except ImportError:
    tqdm = None


SCRIPT_DIR = Path(__file__).resolve().parent
DEFAULT_CHECKPOINT = SCRIPT_DIR / "Task 1_Urban Scene Classification.pth"
DEFAULT_CLASS_MAP = SCRIPT_DIR / "Task 1_Class_to_Index.xlsx"
IMAGE_SUFFIXES = {".jpg", ".jpeg", ".png", ".bmp", ".webp"}


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Run Urban-ImageNet Task 1 HUSIC scene classification."
    )
    parser.add_argument(
        "--input",
        required=True,
        type=Path,
        help="Image file or folder containing images to classify.",
    )
    parser.add_argument(
        "--output",
        default=Path("urban_imagenet_task1_predictions.csv"),
        type=Path,
        help="Output .csv or .xlsx file.",
    )
    parser.add_argument(
        "--checkpoint",
        default=DEFAULT_CHECKPOINT,
        type=Path,
        help="Path to the Task 1 .pth checkpoint.",
    )
    parser.add_argument(
        "--class-map",
        default=DEFAULT_CLASS_MAP,
        type=Path,
        help="Path to the class-index mapping .xlsx file.",
    )
    parser.add_argument("--batch-size", default=32, type=int)
    parser.add_argument("--image-size", default=224, type=int)
    parser.add_argument("--top-k", default=3, type=int)
    parser.add_argument("--recursive", action="store_true")
    parser.add_argument(
        "--device",
        default="auto",
        help="auto, cpu, cuda, or a torch device string.",
    )
    return parser.parse_args()


def get_device(device_arg: str) -> torch.device:
    if device_arg == "auto":
        return torch.device("cuda" if torch.cuda.is_available() else "cpu")
    return torch.device(device_arg)


def read_class_mapping(path: Path) -> dict[int, str]:
    if not path.exists():
        raise FileNotFoundError(f"Class mapping file not found: {path}")

    try:
        import pandas as pd

        df = pd.read_excel(path)
        rows = df[["Label_Index", "Class_Name"]].dropna().itertuples(index=False)
        return {int(label_id): str(class_name) for label_id, class_name in rows}
    except ImportError:
        import openpyxl

        workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
        sheet = workbook.active
        header = [str(cell or "").strip() for cell in next(sheet.iter_rows(values_only=True))]
        label_col = header.index("Label_Index")
        name_col = header.index("Class_Name")
        mapping: dict[int, str] = {}
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if len(row) <= max(label_col, name_col):
                continue
            label_id = row[label_col]
            class_name = row[name_col]
            if label_id is None or class_name is None:
                continue
            mapping[int(label_id)] = str(class_name)
        return mapping


def build_model(num_classes: int) -> nn.Module:
    model = models.resnet152(weights=None)
    in_features = model.fc.in_features
    model.fc = nn.Sequential(
        nn.Linear(in_features, 256),
        nn.ReLU(),
        nn.Dropout(0.5),
        nn.Linear(256, num_classes),
    )
    return model


def load_checkpoint(model: nn.Module, checkpoint_path: Path, device: torch.device) -> nn.Module:
    if not checkpoint_path.exists():
        raise FileNotFoundError(f"Checkpoint not found: {checkpoint_path}")

    checkpoint = torch.load(checkpoint_path, map_location=device)
    if isinstance(checkpoint, dict):
        if "state_dict" in checkpoint:
            checkpoint = checkpoint["state_dict"]
        elif "model_state_dict" in checkpoint:
            checkpoint = checkpoint["model_state_dict"]

    if not isinstance(checkpoint, dict):
        raise ValueError("Unsupported checkpoint format. Expected a PyTorch state_dict.")

    state_dict = {key.replace("module.", ""): value for key, value in checkpoint.items()}
    model.load_state_dict(state_dict)
    model.to(device)
    model.eval()
    return model


def build_transform(image_size: int) -> transforms.Compose:
    return transforms.Compose(
        [
            transforms.Resize((image_size, image_size)),
            transforms.ToTensor(),
            transforms.Normalize(mean=[0.485, 0.456, 0.406], std=[0.229, 0.224, 0.225]),
        ]
    )


def collect_images(input_path: Path, recursive: bool) -> list[Path]:
    if input_path.is_file():
        if input_path.suffix.lower() not in IMAGE_SUFFIXES:
            raise ValueError(f"Unsupported image file type: {input_path.suffix}")
        return [input_path]

    if not input_path.is_dir():
        raise FileNotFoundError(f"Input path not found: {input_path}")

    iterator = input_path.rglob("*") if recursive else input_path.glob("*")
    images = sorted(path for path in iterator if path.suffix.lower() in IMAGE_SUFFIXES)
    if not images:
        raise FileNotFoundError(f"No supported images found under: {input_path}")
    return images


def batched(items: list[Path], batch_size: int) -> Iterable[list[Path]]:
    for start in range(0, len(items), batch_size):
        yield items[start : start + batch_size]


def load_image(path: Path, transform: transforms.Compose) -> torch.Tensor:
    with Image.open(path) as image:
        image = image.convert("RGB")
        return transform(image)


@torch.inference_mode()
def predict(
    model: nn.Module,
    image_paths: list[Path],
    class_mapping: dict[int, str],
    transform: transforms.Compose,
    device: torch.device,
    batch_size: int,
    top_k: int,
) -> list[dict[str, object]]:
    rows: list[dict[str, object]] = []
    batches = list(batched(image_paths, batch_size))
    progress = tqdm(batches, desc="Predicting") if tqdm else batches

    for batch_paths in progress:
        tensors: list[torch.Tensor] = []
        valid_paths: list[Path] = []
        for path in batch_paths:
            try:
                tensors.append(load_image(path, transform))
                valid_paths.append(path)
            except Exception as exc:
                rows.append(
                    {
                        "filename": path.name,
                        "path": str(path),
                        "error": str(exc),
                    }
                )

        if not tensors:
            continue

        inputs = torch.stack(tensors).to(device)
        probabilities = torch.softmax(model(inputs), dim=1)
        values, indices = torch.topk(probabilities, k=min(top_k, probabilities.shape[1]), dim=1)

        for path, probs, labels in zip(valid_paths, values.cpu(), indices.cpu()):
            top_items = [
                {
                    "index": int(label),
                    "class_name": class_mapping.get(int(label), str(int(label))),
                    "probability": float(prob),
                }
                for prob, label in zip(probs, labels)
            ]
            best = top_items[0]
            row: dict[str, object] = {
                "filename": path.name,
                "path": str(path),
                "predicted_index": best["index"],
                "predicted_class": best["class_name"],
                "confidence": best["probability"],
            }
            for rank, item in enumerate(top_items, start=1):
                row[f"top{rank}_index"] = item["index"]
                row[f"top{rank}_class"] = item["class_name"]
                row[f"top{rank}_probability"] = item["probability"]
            rows.append(row)

    return rows


def write_rows(rows: list[dict[str, object]], output_path: Path) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    fieldnames = sorted({key for row in rows for key in row.keys()})

    if output_path.suffix.lower() in {".xlsx", ".xls"}:
        try:
            import pandas as pd
        except ImportError as exc:
            raise RuntimeError("Install pandas and openpyxl to write Excel output.") from exc
        pd.DataFrame(rows).to_excel(output_path, index=False)
        return

    with output_path.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


def main() -> None:
    args = parse_args()
    device = get_device(args.device)
    class_mapping = read_class_mapping(args.class_map)
    image_paths = collect_images(args.input, args.recursive)
    transform = build_transform(args.image_size)

    model = build_model(num_classes=len(class_mapping))
    model = load_checkpoint(model, args.checkpoint, device)

    rows = predict(
        model=model,
        image_paths=image_paths,
        class_mapping=class_mapping,
        transform=transform,
        device=device,
        batch_size=args.batch_size,
        top_k=args.top_k,
    )
    write_rows(rows, args.output)

    print(f"Device: {device}")
    print(f"Images: {len(image_paths)}")
    print(f"Output: {args.output}")


if __name__ == "__main__":
    main()
